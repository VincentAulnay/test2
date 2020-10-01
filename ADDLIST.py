from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlwt
import xlrd
import time
from xlutils.copy import copy
import datetime
import datetime as dt
from tkinter import filedialog
from tkinter import *
from bs4 import BeautifulSoup
from urllib.request import urlopen
import os
import shutil
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
import threading
import sys


wbx = load_workbook(path_RESULT.filename)
ws = wbx.active
nrow=ws.max_row
print(nrow)
listnewann=['https://www.airbnb.fr/rooms/11292061', 'https://www.airbnb.fr/rooms/39656481', 'https://www.airbnb.fr/rooms/43984593', 'https://www.airbnb.fr/rooms/43882152', 'https://www.airbnb.fr/rooms/42035404', 'https://www.airbnb.fr/rooms/37063937', 'https://www.airbnb.fr/rooms/43751234', 'https://www.airbnb.fr/rooms/35079536', 'https://www.airbnb.fr/rooms/38118668', 'https://www.airbnb.fr/rooms/39164302', 'https://www.airbnb.fr/rooms/35572921', 'https://www.airbnb.fr/rooms/45094512', 'https://www.airbnb.fr/rooms/41658404', 'https://www.airbnb.fr/rooms/12804992', 'https://www.airbnb.fr/rooms/37790351', 'https://www.airbnb.fr/rooms/41207520', 'https://www.airbnb.fr/rooms/43703883', 'https://www.airbnb.fr/rooms/45174079', 'https://www.airbnb.fr/rooms/44879732', 'https://www.airbnb.fr/rooms/44875175', 'https://www.airbnb.fr/rooms/41659298', 'https://www.airbnb.fr/rooms/44878281', 'https://www.airbnb.fr/rooms/42920627', 'https://www.airbnb.fr/rooms/44880185', 'https://www.airbnb.fr/rooms/44880465', 'https://www.airbnb.fr/rooms/24739003', 'https://www.airbnb.fr/rooms/43895363', 'https://www.airbnb.fr/rooms/38715037', 'https://www.airbnb.fr/rooms/35091157', 'https://www.airbnb.fr/rooms/39751245', 'https://www.airbnb.fr/rooms/3598598', 'https://www.airbnb.fr/rooms/38125620', 'https://www.airbnb.fr/rooms/38499588', 'https://www.airbnb.fr/rooms/3599496', 'https://www.airbnb.fr/rooms/42557582', 'https://www.airbnb.fr/rooms/38124111', 'https://www.airbnb.fr/rooms/37356268', 'https://www.airbnb.fr/rooms/35593179', 'https://www.airbnb.fr/rooms/29454994', 'https://www.airbnb.fr/rooms/39813747', 'https://www.airbnb.fr/rooms/20612833', 'https://www.airbnb.fr/rooms/38886988', 'https://www.airbnb.fr/rooms/44165828', 'https://www.airbnb.fr/rooms/plus/33443645', 'https://www.airbnb.fr/rooms/42174532', 'https://www.airbnb.fr/rooms/38986489', 'https://www.airbnb.fr/rooms/41393526', 'https://www.airbnb.fr/rooms/37311127', 'https://www.airbnb.fr/rooms/44416257', 'https://www.airbnb.fr/rooms/43796044', 'https://www.airbnb.fr/rooms/44220264', 'https://www.airbnb.fr/rooms/40607646', 'https://www.airbnb.fr/rooms/37581737', 'https://www.airbnb.fr/rooms/35983895', 'https://www.airbnb.fr/rooms/31758120', 'https://www.airbnb.fr/rooms/40230988', 'https://www.airbnb.fr/rooms/31372696', 'https://www.airbnb.fr/rooms/12936935', 'https://www.airbnb.fr/rooms/42162970', 'https://www.airbnb.fr/rooms/41673815', 'https://www.airbnb.fr/rooms/43289208', 'https://www.airbnb.fr/rooms/36473847', 'https://www.airbnb.fr/rooms/44301226', 'https://www.airbnb.fr/rooms/36382616', 'https://www.airbnb.fr/rooms/38087288', 'https://www.airbnb.fr/rooms/43876760', 'https://www.airbnb.fr/rooms/39163467', 'https://www.airbnb.fr/rooms/37377617', 'https://www.airbnb.fr/rooms/13146633', 'https://www.airbnb.fr/rooms/27267240', 'https://www.airbnb.fr/rooms/43979118', 'https://www.airbnb.fr/rooms/41498087', 'https://www.airbnb.fr/rooms/35783400', 'https://www.airbnb.fr/rooms/44210780', 'https://www.airbnb.fr/rooms/42034241', 'https://www.airbnb.fr/rooms/42593676', 'https://www.airbnb.fr/rooms/24672229', 'https://www.airbnb.fr/rooms/45433041']
i=1
for h in listnewann:
  ws.cell(row=2, column=nrow+i).value=h
  i=i+1
wbx.save(path_RESULT.filename)
wbx = load_workbook(path_RESULT.filename)
ws = wbx.active
nrow=ws.max_row
print(nrow)
